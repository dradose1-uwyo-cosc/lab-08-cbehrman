#Caleb Behrman
#lab section 15
#11/19/24

import openpyxl
from openpyxl.styles import PatternFill

#this code below creates the new workbook
wb=openpyxl.Workbook()
sheet=wb.active

#im setting the area/size/square stuff here
column_width=2 
row_height=12  

#this code below makes it to whree each colum has the same width
for col in sheet.columns:
    sheet.column_dimensions[col[0].column_letter].width=column_width

#same thing here but for the height
for row in sheet.iter_rows():
    sheet.row_dimensions[row[0].row].height=row_height

#defines the colors bwlow
colors={
    'red': 'FFFF0000',  
    'green': 'FF00FF00',  
    'blue': 'FF0000FF',  
    'yellow': 'FFFFFF00',  
    'black': 'FF000000' 
}

#this is the pattern, I just repeated the colors to make a cool design. Its not like a charector or anything like you showed us in the description of this assingment
pixel_art=[
    ['red', 'green', 'blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black'],
    ['green', 'blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black', 'red'],
    ['blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black', 'red', 'green'],
    ['yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black', 'red', 'green', 'blue'],
    ['black', 'red', 'green', 'blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow'],
    ['red', 'green', 'blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black'],
    ['green', 'blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black', 'red'],
    ['blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black', 'red', 'green'],
    ['yellow', 'black', 'red', 'green', 'blue', 'yellow', 'black', 'red', 'green', 'blue'],
    ['black', 'red', 'green', 'blue', 'yellow', 'black', 'red', 'green', 'blue', 'yellow']
]

#this just loops through the grid and applys the colors i guess you could say
for row_idx, row in enumerate(pixel_art, start=1):
    for col_idx, color in enumerate(row, start=1):
        #fills colors
        fill=PatternFill(start_color=colors[color], end_color=colors[color], fill_type="solid")
        cell=sheet.cell(row=row_idx, column=col_idx)
        cell.fill=fill

#this just saves it to a file
wb.save("pixel_art_10x10.xlsx")

print("Pixel art has been generated and saved as 'pixel_art_10x10.xlsx'.")